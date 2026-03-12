from io import BytesIO
from openpyxl import Workbook


def generate_excel(top20, noun_top10, verb_top10, adj_top10, adv_top10):
    wb = Workbook()

    # 第1个sheet：总词频前20
    ws = wb.active
    ws.title = "总词频Top20"
    ws.append(["排名", "词汇", "频次"])
    for i, item in enumerate(top20, start=1):
        ws.append([i, item["token"], item["count"]])

    # 第2个sheet：名词Top10
    ws2 = wb.create_sheet("名词Top10")
    ws2.append(["排名", "词汇", "频次"])
    for i, item in enumerate(noun_top10, start=1):
        ws2.append([i, item["token"], item["count"]])

    # 第3个sheet：动词Top10
    ws3 = wb.create_sheet("动词Top10")
    ws3.append(["排名", "词汇", "频次"])
    for i, item in enumerate(verb_top10, start=1):
        ws3.append([i, item["token"], item["count"]])

    # 第4个sheet：形容词Top10
    ws4 = wb.create_sheet("形容词Top10")
    ws4.append(["排名", "词汇", "频次"])
    for i, item in enumerate(adj_top10, start=1):
        ws4.append([i, item["token"], item["count"]])

    # 第5个sheet：副词Top10
    ws5 = wb.create_sheet("副词Top10")
    ws5.append(["排名", "词汇", "频次"])
    for i, item in enumerate(adv_top10, start=1):
        ws5.append([i, item["token"], item["count"]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output